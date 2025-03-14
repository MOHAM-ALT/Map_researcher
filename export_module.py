#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Export Module for Map_researcher 0.4
This module provides functionality for exporting hotel data and reports.
"""

import os
import sys
import time
import json
import logging
import csv
from typing import Dict, List, Any, Tuple, Optional, Union
from datetime import datetime, timedelta

# Initialize logger
logger = logging.getLogger("export_module")

class ExportModule:
    """Class for data and report export operations"""
    
    def __init__(self, db=None, scraper=None):
        """
        Initialize the export module
        
        Args:
            db: Database instance
            scraper: HotelScraper instance
        """
        self.db = db
        self.scraper = scraper
    
    def export_excel(self, data: List[Dict] = None, output_path: str = None, 
                    format_type: str = 'detailed', query: Dict = None) -> Dict:
        """
        Export data to Excel format
        
        Args:
            data: Data to export (if None, will use query to get data)
            output_path: Output file path
            format_type: Export format type (simple, detailed, multi_sheet)
            query: Query to use if data not provided
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting data to Excel: {output_path}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Get data if not provided
        if not data and self.db:
            try:
                data = self.db.search_hotels(query or {})
                logger.info(f"Retrieved {len(data)} hotels for export")
            except Exception as e:
                logger.error(f"Error retrieving data: {e}")
                return {"error": f"Database error: {str(e)}"}
        
        if not data:
            logger.error("No data to export")
            return {"error": "No data to export"}
        
        try:
            # Import pandas and openpyxl for Excel export
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = Workbook()
            
            if format_type == 'simple':
                # Simple format with basic fields
                ws = wb.active
                ws.title = "Hotels"
                
                # Define fields to export
                fields = ['id', 'name', 'address', 'city', 'country', 'stars', 'data_source']
                
                # Create DataFrame with selected fields
                df = pd.DataFrame([{k: h.get(k, '') for k in fields} for h in data])
                
                # Add data to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Format header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            elif format_type == 'detailed':
                # Detailed format with all fields in one sheet
                ws = wb.active
                ws.title = "Hotels Detailed"
                
                # Add all fields
                df = pd.DataFrame(data)
                
                # Add data to worksheet
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # Format header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            else:  # multi_sheet
                # Multiple sheets for different aspects
                # Sheet 1: Basic Info
                ws_basic = wb.active
                ws_basic.title = "Basic Info"
                
                basic_fields = ['id', 'name', 'address', 'city', 'country', 'stars', 'price_range', 'last_updated']
                df_basic = pd.DataFrame([{k: h.get(k, '') for k in basic_fields} for h in data])
                
                for r in dataframe_to_rows(df_basic, index=False, header=True):
                    ws_basic.append(r)
                
                # Format header
                for cell in ws_basic[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Sheet 2: Contact Info
                ws_contact = wb.create_sheet("Contact Info")
                
                contact_fields = ['id', 'name', 'phone', 'email', 'website']
                df_contact = pd.DataFrame([{k: h.get(k, '') for k in contact_fields} for h in data])
                
                for r in dataframe_to_rows(df_contact, index=False, header=True):
                    ws_contact.append(r)
                
                # Format header
                for cell in ws_contact[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Sheet 3: Facilities
                ws_facilities = wb.create_sheet("Facilities")
                
                facilities_fields = ['id', 'name', 'facilities', 'stars', 'price_range']
                df_facilities = pd.DataFrame([{k: h.get(k, '') for k in facilities_fields} for h in data])
                
                for r in dataframe_to_rows(df_facilities, index=False, header=True):
                    ws_facilities.append(r)
                
                # Format header
                for cell in ws_facilities[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Sheet 4: Risk Analysis (if available)
                risk_data = []
                for hotel in data:
                    if hotel.get('risk_analysis') or hotel.get('risk_score'):
                        risk_level = (hotel.get('risk_analysis', {}).get('risk_level') or 
                                     'high' if hotel.get('risk_score', 0) >= 7 else 
                                     'medium' if hotel.get('risk_score', 0) >= 3 else 'low')
                        
                        risk_data.append({
                            'id': hotel.get('id', ''),
                            'name': hotel.get('name', ''),
                            'risk_score': hotel.get('risk_score', hotel.get('risk_analysis', {}).get('risk_score', 0)),
                            'risk_level': risk_level,
                            'risk_factors': str(hotel.get('risk_analysis', {}).get('risk_factors', []))
                        })
                
                if risk_data:
                    ws_risk = wb.create_sheet("Risk Analysis")
                    df_risk = pd.DataFrame(risk_data)
                    
                    for r in dataframe_to_rows(df_risk, index=False, header=True):
                        ws_risk.append(r)
                    
                    # Format header
                    for cell in ws_risk[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Save workbook
            wb.save(output_path)
            
            logger.info(f"Excel export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Data exported successfully to {output_path}",
                "file_path": output_path,
                "record_count": len(data)
            }
        
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}
    
    def export_csv(self, data: List[Dict] = None, output_path: str = None, 
                  format_type: str = 'detailed', query: Dict = None) -> Dict:
        """
        Export data to CSV format
        
        Args:
            data: Data to export (if None, will use query to get data)
            output_path: Output file path
            format_type: Export format type (simple, detailed, multiple)
            query: Query to use if data not provided
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting data to CSV: {output_path}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        # Get data if not provided
        if not data and self.db:
            try:
                data = self.db.search_hotels(query or {})
                logger.info(f"Retrieved {len(data)} hotels for export")
            except Exception as e:
                logger.error(f"Error retrieving data: {e}")
                return {"error": f"Database error: {str(e)}"}
        
        if not data:
            logger.error("No data to export")
            return {"error": "No data to export"}
        
        try:
            # Handle different format types
            if format_type == 'simple':
                # Simple format with basic fields
                fields = ['id', 'name', 'address', 'city', 'country', 'stars', 'data_source']
                
                # Create CSV file
                with open(output_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
                    writer.writeheader()
                    for hotel in data:
                        writer.writerow({k: hotel.get(k, '') for k in fields})
            
            elif format_type == 'detailed':
                # Detailed format with all fields
                # Get all possible field names
                all_fields = set()
                for hotel in data:
                    all_fields.update(hotel.keys())
                
                # Sort fields for consistent output
                sorted_fields = sorted(all_fields)
                
                # Create CSV file
                with open(output_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=sorted_fields, extrasaction='ignore')
                    writer.writeheader()
                    for hotel in data:
                        writer.writerow(hotel)
            
            else:  # 'multiple'
                # Multiple CSV files for different aspects
                # Get base name without extension
                base_path = os.path.splitext(output_path)[0]
                
                # Basic info CSV
                basic_fields = ['id', 'name', 'address', 'city', 'country', 'stars', 'price_range', 'last_updated']
                basic_path = f"{base_path}_basic.csv"
                
                with open(basic_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=basic_fields, extrasaction='ignore')
                    writer.writeheader()
                    for hotel in data:
                        writer.writerow({k: hotel.get(k, '') for k in basic_fields})
                
                # Contact info CSV
                contact_fields = ['id', 'name', 'phone', 'email', 'website']
                contact_path = f"{base_path}_contact.csv"
                
                with open(contact_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=contact_fields, extrasaction='ignore')
                    writer.writeheader()
                    for hotel in data:
                        writer.writerow({k: hotel.get(k, '') for k in contact_fields})
                
                # Set output path to the base path for result
                output_path = base_path
            
            logger.info(f"CSV export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Data exported successfully to {output_path}",
                "file_path": output_path,
                "record_count": len(data)
            }
        
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}
    
    def export_json(self, data: List[Dict] = None, output_path: str = None, 
                   format_type: str = 'detailed', query: Dict = None) -> Dict:
        """
        Export data to JSON format
        
        Args:
            data: Data to export (if None, will use query to get data)
            output_path: Output file path
            format_type: Export format type (simple, detailed)
            query: Query to use if data not provided
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting data to JSON: {output_path}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/hotels_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        # Get data if not provided
        if not data and self.db:
            try:
                data = self.db.search_hotels(query or {})
                logger.info(f"Retrieved {len(data)} hotels for export")
            except Exception as e:
                logger.error(f"Error retrieving data: {e}")
                return {"error": f"Database error: {str(e)}"}
        
        if not data:
            logger.error("No data to export")
            return {"error": "No data to export"}
        
        try:
            # Handle different format types
            if format_type == 'simple':
                # Simple format with basic fields
                fields = ['id', 'name', 'address', 'city', 'country', 'stars', 'data_source']
                
                # Filter data to include only basic fields
                filtered_data = [{k: h.get(k, '') for k in fields} for h in data]
                
                # Export to JSON
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(filtered_data, f, ensure_ascii=False, indent=2)
            
            else:  # 'detailed'
                # Export full data
                with open(output_path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"JSON export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Data exported successfully to {output_path}",
                "file_path": output_path,
                "record_count": len(data)
            }
        
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}
    
    def export_map(self, data: List[Dict] = None, output_path: str = None, 
                  query: Dict = None, map_title: str = None) -> Dict:
        """
        Export hotel data to interactive map (HTML)
        
        Args:
            data: Data to export (if None, will use query to get data)
            output_path: Output file path
            query: Query to use if data not provided
            map_title: Title for the map
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting data to map: {output_path}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/hotels_map_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        # Get data if not provided
        if not data and self.db:
            try:
                data = self.db.search_hotels(query or {})
                logger.info(f"Retrieved {len(data)} hotels for map export")
            except Exception as e:
                logger.error(f"Error retrieving data: {e}")
                return {"error": f"Database error: {str(e)}"}
        
        if not data:
            logger.error("No data to export")
            return {"error": "No data to export"}
        
        try:
            # Import folium for map creation
            import folium
            from folium.plugins import MarkerCluster
            
            # Filter hotels with coordinates
            hotels_with_coords = [h for h in data if h.get('latitude') and h.get('longitude')]
            
            if not hotels_with_coords:
                logger.error("No hotels with coordinates found")
                return {"error": "No hotels with coordinates found"}
            
            # Calculate map center (average of coordinates)
            lat_sum = sum(float(h.get('latitude', 0)) for h in hotels_with_coords)
            lng_sum = sum(float(h.get('longitude', 0)) for h in hotels_with_coords)
            center_lat = lat_sum / len(hotels_with_coords)
            center_lng = lng_sum / len(hotels_with_coords)
            
            # Create map
            title = map_title or f"Hotel Map ({len(hotels_with_coords)} hotels)"
            hotel_map = folium.Map(location=[center_lat, center_lng], zoom_start=13, tiles="OpenStreetMap")
            
            # Add title
            title_html = f'''
                <h3 align="center" style="font-size:16px"><b>{title}</b></h3>
                <p align="center">Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            '''
            hotel_map.get_root().html.add_child(folium.Element(title_html))
            
            # Use marker cluster for better performance with many markers
            marker_cluster = MarkerCluster().add_to(hotel_map)
            
            # Add markers for each hotel
            for hotel in hotels_with_coords:
                # Prepare tooltip content
                tooltip = f"{hotel.get('name', 'Unknown Hotel')}"
                
                # Prepare popup content
                popup_content = f"""
                <div style="min-width:200px">
                    <h4>{hotel.get('name', 'Unknown Hotel')}</h4>
                    <p><b>Address:</b> {hotel.get('address', 'N/A')}</p>
                """
                
                if hotel.get('stars'):
                    popup_content += f"<p><b>Rating:</b> {hotel.get('stars')} ★</p>"
                
                if hotel.get('phone'):
                    popup_content += f"<p><b>Phone:</b> {hotel.get('phone')}</p>"
                
                if hotel.get('website'):
                    popup_content += f'<p><b>Website:</b> <a href="{hotel.get("website")}" target="_blank">{hotel.get("website")}</a></p>'
                
                popup_content += f"<p><b>Source:</b> {hotel.get('data_source', 'N/A')}</p>"
                
                # Add risk info if available
                if hotel.get('risk_level') or hotel.get('risk_score') or hotel.get('risk_analysis'):
                    risk_level = (hotel.get('risk_level') or 
                                hotel.get('risk_analysis', {}).get('risk_level') or 
                                ('high' if hotel.get('risk_score', 0) >= 7 else 
                                'medium' if hotel.get('risk_score', 0) >= 3 else 'low'))
                    
                    risk_color = {'high': 'red', 'medium': 'orange', 'low': 'green'}.get(risk_level, 'gray')
                    
                    popup_content += f'<p><b>Risk Level:</b> <span style="color:{risk_color};font-weight:bold;">{risk_level.upper()}</span></p>'
                
                popup_content += "</div>"
                
                # Set icon color based on data source
                icon_color = 'blue'
                if hotel.get('data_source') == 'Google Places':
                    icon_color = 'red'
                elif hotel.get('data_source') == 'OpenStreetMap':
                    icon_color = 'green'
                elif hotel.get('risk_level') == 'high' or (hotel.get('risk_score', 0) >= 7):
                    icon_color = 'darkred'
                elif hotel.get('risk_level') == 'medium' or (hotel.get('risk_score', 0) >= 3):
                    icon_color = 'orange'
                
                # Create marker
                folium.Marker(
                    location=[float(hotel.get('latitude')), float(hotel.get('longitude'))],
                    popup=folium.Popup(popup_content, max_width=300),
                    tooltip=tooltip,
                    icon=folium.Icon(color=icon_color, icon='hotel', prefix='fa')
                ).add_to(marker_cluster)
            
            # Save map to file
            hotel_map.save(output_path)
            
            logger.info(f"Map export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Map exported successfully to {output_path}",
                "file_path": output_path,
                "record_count": len(hotels_with_coords)
            }
        
        except Exception as e:
            logger.error(f"Error exporting to map: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}
    
    def export_violations_report(self, data: Dict = None, output_path: str = None, 
                                format_type: str = 'detailed') -> Dict:
        """
        Export violations report
        
        Args:
            data: Report data (if None, will generate new report)
            output_path: Output file path
            format_type: Export format type (simple, detailed)
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting violations report: {output_path}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/violations_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        # Generate report if not provided
        if not data and hasattr(self, 'violation_detection') and self.violation_detection:
            try:
                data = self.violation_detection.create_violations_report()
                logger.info("Generated new violations report")
            except Exception as e:
                logger.error(f"Error generating violations report: {e}")
                return {"error": f"Report generation failed: {str(e)}"}
        
        if not data:
            logger.error("No report data to export")
            return {"error": "No report data to export"}
        
        try:
            # Generate HTML report
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <meta name="viewport" content="width=device-width, initial-scale=1">
                <title>Hotel Violations Report</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; }}
                    h1, h2, h3 {{ color: #333; }}
                    .container {{ max-width: 1200px; margin: 0 auto; }}
                    .header {{ background-color: #f2f2f2; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
                    .section {{ margin-bottom: 30px; }}
                    table {{ border-collapse: collapse; width: 100%; }}
                    th, td {{ border: 1px solid #ddd; padding: 8px; }}
                    th {{ background-color: #f2f2f2; text-align: left; }}
                    tr:nth-child(even) {{ background-color: #f9f9f9; }}
                    .high {{ color: #d9534f; }}
                    .medium {{ color: #f0ad4e; }}
                    .low {{ color: #5cb85c; }}
                    .summary-box {{ display: inline-block; width: 200px; height: 100px; margin: 10px;
                                  padding: 15px; border-radius: 5px; text-align: center; }}
                    .high-bg {{ background-color: #ffebee; }}
                    .medium-bg {{ background-color: #fff8e1; }}
                    .low-bg {{ background-color: #e8f5e9; }}
                    .recommendations {{ background-color: #e3f2fd; padding: 15px; border-radius: 5px; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Hotel Violations Report</h1>
                        <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    </div>
                    
                    <div class="section">
                        <h2>Summary</h2>
                        <div class="summary-box high-bg">
                            <h3>High Risk</h3>
                            <p>{data.get('summary', {}).get('high_risk_count', 0)}</p>
                        </div>
                        <div class="summary-box medium-bg">
                            <h3>Medium Risk</h3>
                            <p>{data.get('summary', {}).get('medium_risk_count', 0)}</p>
                        </div>
                        <div class="summary-box low-bg">
                            <h3>Low Risk</h3>
                            <p>{data.get('summary', {}).get('low_risk_count', 0)}</p>
                        </div>
                        <p>Total Hotels Analyzed: {data.get('summary', {}).get('total_hotels', 0)}</p>
                    </div>
            """
            
            # Add risk factors section
            html_content += """
                    <div class="section">
                        <h2>Risk Factors</h2>
                        <table>
                            <tr>
                                <th>Risk Factor</th>
                                <th>Count</th>
                            </tr>
            """
            
            # Add risk factors
            for factor, count in data.get('risk_factors', {}).items():
                html_content += f"""
                            <tr>
                                <td>{factor}</td>
                                <td>{count}</td>
                            </tr>
                """
            
            html_content += """
                        </table>
                    </div>
            """
            
            # Add recommendations section
            html_content += """
                    <div class="section recommendations">
                        <h2>Recommendations</h2>
                        <ul>
            """
            
            # Add recommendations
            for rec in data.get('recommendations', []):
                priority_class = rec.get('priority', 'medium')
                html_content += f"""
                            <li class="{priority_class}">{rec.get('description', '')}</li>
                """
            
            html_content += """
                        </ul>
                    </div>
            """
            
            # Add high risk hotels section if format_type is detailed
            if format_type == 'detailed' and data.get('high_risk_hotels'):
                html_content += """
                    <div class="section">
                        <h2>High Risk Hotels</h2>
                        <table>
                            <tr>
                                <th>ID</th>
                                <th>Name</th>
                                <th>Address</th>
                                <th>Risk Score</th>
                                <th>Risk Factors</th>
                            </tr>
                """
                
                # Add high risk hotels
                for hotel in data.get('high_risk_hotels', []):
                    risk_factors_text = ""
                    for factor in hotel.get('risk_analysis', {}).get('risk_factors', []):
                        risk_factors_text += f"<li>{factor.get('details', '')}</li>"
                    
                    html_content += f"""
                            <tr>
                                <td>{hotel.get('id', '')}</td>
                                <td>{hotel.get('name', '')}</td>
                                <td>{hotel.get('address', '')}</td>
                                <td class="high">{hotel.get('risk_analysis', {}).get('risk_score', 0)}</td>
                                <td><ul>{risk_factors_text}</ul></td>
                            </tr>
                    """
                
                html_content += """
                        </table>
                    </div>
                """
            
            # Close HTML
            html_content += """
                </div>
            </body>
            </html>
            """
            
            # Write to file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"Violations report export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Violations report exported successfully to {output_path}",
                "file_path": output_path
            }
        
        except Exception as e:
            logger.error(f"Error exporting violations report: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}
    
    def export_timeline(self, hotel_id: str = None, output_path: str = None, 
                       hotel_name: str = None) -> Dict:
        """
        Export timeline chart for a hotel
        
        Args:
            hotel_id: Hotel ID
            output_path: Output file path
            hotel_name: Hotel name (for display purposes)
            
        Returns:
            Dict: Export result
        """
        logger.info(f"Exporting timeline chart for hotel: {hotel_id}")
        
        # Set default output path if not provided
        if not output_path:
            os.makedirs("exports", exist_ok=True)
            output_path = f"exports/hotel_timeline_{hotel_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
        
        # Get hotel data
        hotel_data = None
        if hotel_id and self.db and hasattr(self.db, 'get_hotel_by_id'):
            try:
                hotel_data = self.db.get_hotel_by_id(hotel_id)
                if not hotel_name and hotel_data:
                    hotel_name = hotel_data.get('name', f"Hotel ID: {hotel_id}")
            except Exception as e:
                logger.error(f"Error retrieving hotel data: {e}")
                return {"error": f"Database error: {str(e)}"}
        
        if not hotel_data:
            logger.error("No hotel data found")
            return {"error": "Hotel not found"}
        
        # Get hotel history
        history = []
        if hotel_id and self.db and hasattr(self.db, 'get_hotel_history'):
            try:
                history = self.db.get_hotel_history(hotel_id)
                logger.info(f"Retrieved {len(history)} history records for hotel")
            except Exception as e:
                logger.error(f"Error retrieving hotel history: {e}")
                return {"error": f"History retrieval failed: {str(e)}"}
        
        if not history:
            logger.error("No history records found")
            return {"error": "No history records found for this hotel"}
        
        try:
            # Generate HTML timeline
            html_content = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="utf-8">
                <meta name="viewport" content="width=device-width, initial-scale=1">
                <title>Hotel Timeline - {hotel_name}</title>
                <style>
                    body {{ font-family: Arial, sans-serif; margin: 0; padding: 20px; }}
                    h1, h2, h3 {{ color: #333; }}
                    .container {{ max-width: 1200px; margin: 0 auto; }}
                    .header {{ background-color: #f2f2f2; padding: 20px; border-radius: 5px; margin-bottom: 20px; }}
                    .timeline {{ position: relative; max-width: 1200px; margin: 0 auto; }}
                    .timeline::after {{ content: ''; position: absolute; width: 6px; background-color: #999; top: 0; bottom: 0; left: 50%; margin-left: -3px; }}
                    .container-left {{ padding: 10px 40px; position: relative; background-color: inherit; width: 45%; left: 0; }}
                    .container-right {{ padding: 10px 40px; position: relative; background-color: inherit; width: 45%; left: 50%; }}
                    .content {{ padding: 20px; background-color: white; position: relative; border-radius: 6px; border: 1px solid #ddd; }}
                    .container-left .content::after {{ content: " "; position: absolute; top: 22px; right: -15px; border-width: 10px 0 10px 15px; border-color: transparent transparent transparent white; border-style: solid; }}
                    .container-right .content::after {{ content: " "; position: absolute; top: 22px; left: -15px; border-width: 10px 15px 10px 0; border-color: transparent white transparent transparent; border-style: solid; }}
                    .name-change {{ background-color: #e3f2fd; }}
                    .ownership-change {{ background-color: #fff8e1; }}
                    .status-change {{ background-color: #ffebee; }}
                    .platform-change {{ background-color: #e8f5e9; }}
                    .date {{ position: relative; color: #666; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="header">
                        <h1>Timeline for {hotel_name}</h1>
                        <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    </div>
                    
                    <div class="timeline">
            """
            
            # Sort history by date
            history.sort(key=lambda x: x.get('event_date', ''))
            
            # Add timeline events
            for i, event in enumerate(history):
                event_date = event.get('event_date', '').split('T')[0] if 'T' in event.get('event_date', '') else event.get('event_date', '')
                event_type = event.get('event_type', '')
                old_value = event.get('old_value', '')
                new_value = event.get('new_value', '')
                
                # Determine event class
                event_class = ""
                if 'name' in event_type.lower():
                    event_class = "name-change"
                elif 'owner' in event_type.lower():
                    event_class = "ownership-change"
                elif 'status' in event_type.lower() or 'classification' in event_type.lower():
                    event_class = "status-change"
                elif 'platform' in event_type.lower() or 'listing' in event_type.lower():
                    event_class = "platform-change"
                
                # Alternate between left and right
                container_class = "container-left" if i % 2 == 0 else "container-right"
                
                html_content += f"""
                        <div class="{container_class}">
                            <div class="date">{event_date}</div>
                            <div class="content {event_class}">
                                <h3>{event_type}</h3>
                                <p>From: {old_value}</p>
                                <p>To: {new_value}</p>
                                <p>Source: {event.get('source', 'N/A')}</p>
                            </div>
                        </div>
                """
            
            # Close HTML
            html_content += """
                    </div>
                </div>
            </body>
            </html>
            """
            
            # Write to file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            logger.info(f"Timeline export completed successfully: {output_path}")
            return {
                "status": "success",
                "message": f"Timeline exported successfully to {output_path}",
                "file_path": output_path,
                "event_count": len(history)
            }
        
        except Exception as e:
            logger.error(f"Error exporting timeline: {e}")
            import traceback
            traceback.print_exc()
            return {"error": f"Export failed: {str(e)}"}

# If run directly, display module info
if __name__ == "__main__":
    print("Export Module for Map_researcher 0.4")
    print("This module provides functionality for exporting hotel data and reports.")